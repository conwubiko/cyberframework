"""Export service for CSV and Excel formats."""
import csv
import io
from app.models.assessment import Assessment, AssessmentResponse
from app.models.scan import ScanJob, ScanFinding
from app.data.framework_controls import CROF_FUNCTIONS
from app.services.scoring import get_assessment_results


def export_assessment_csv(assessment_id):
    """Export assessment responses as CSV."""
    assessment = Assessment.query.get(assessment_id)
    if not assessment:
        return None

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Control ID', 'Function ID', 'Title', 'Answer', 'Score', 'Evidence Notes'])

    control_titles = {}
    for func in CROF_FUNCTIONS:
        for ctrl in func['controls']:
            control_titles[ctrl['id']] = ctrl['title']

    for resp in assessment.responses.all():
        writer.writerow([
            resp.control_id,
            resp.function_id,
            control_titles.get(resp.control_id, ''),
            resp.answer,
            resp.score,
            resp.evidence_notes,
        ])

    return output.getvalue()


def export_assessment_excel(assessment_id):
    """Export assessment responses as Excel."""
    from openpyxl import Workbook

    assessment = Assessment.query.get(assessment_id)
    if not assessment:
        return None

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    responses = [
        {'control_id': r.control_id, 'function_id': r.function_id, 'answer': r.answer}
        for r in assessment.responses.all()
    ]
    results = get_assessment_results(responses)

    ws_summary.append(['Assessment', assessment.title])
    ws_summary.append(['Overall Score', f"{results['overall_score']}%"])
    ws_summary.append(['Maturity Level', results['maturity_level']])
    ws_summary.append(['Status', assessment.status])
    ws_summary.append([])
    ws_summary.append(['Function', 'Score %', 'Answered', 'Total'])
    for fid, fs in results['function_scores'].items():
        ws_summary.append([fs['name'], fs['score_pct'], fs['answered'], fs['total']])

    # Responses sheet
    ws_resp = wb.create_sheet('Responses')
    ws_resp.append(['Control ID', 'Function ID', 'Title', 'Answer', 'Score', 'Evidence Notes'])

    control_titles = {}
    for func in CROF_FUNCTIONS:
        for ctrl in func['controls']:
            control_titles[ctrl['id']] = ctrl['title']

    for resp in assessment.responses.all():
        ws_resp.append([
            resp.control_id,
            resp.function_id,
            control_titles.get(resp.control_id, ''),
            resp.answer,
            resp.score,
            resp.evidence_notes or '',
        ])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_findings_csv(scan_id):
    """Export scan findings as CSV."""
    scan = ScanJob.query.get(scan_id)
    if not scan:
        return None

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Severity', 'Control ID', 'Title', 'Description', 'Remediation'])

    for f in scan.findings.all():
        writer.writerow([f.severity, f.control_id, f.title, f.description, f.remediation])

    return output.getvalue()


def export_findings_excel(scan_id):
    """Export scan findings as Excel."""
    from openpyxl import Workbook

    scan = ScanJob.query.get(scan_id)
    if not scan:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = 'Findings'

    ws.append(['Scan Target', scan.target])
    ws.append(['Scan Type', scan.scan_type])
    ws.append(['Status', scan.status])
    ws.append([])
    ws.append(['Severity', 'Control ID', 'Title', 'Description', 'Remediation'])

    for f in scan.findings.all():
        ws.append([f.severity, f.control_id, f.title, f.description, f.remediation])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
